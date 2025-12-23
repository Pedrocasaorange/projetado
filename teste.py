# sistema_fluxo_caixa_completo.py - Sistema Completo de Fluxo de Caixa
from flask import Flask, render_template_string, request, jsonify, redirect, url_for, session
from functools import wraps
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
app.secret_key = 'Chave_segurança_projetado_2025'

# Configurações
DATA_FILE = 'fluxo_caixa_dados.json'

# Dados dos empreendimentos
EMPREENDIMENTOS = [
    "ALLEGRO", "PIAZZA", "CASA PARQUE", "CASA BOA VIAGEM", 
    "CASA MAYOR", "CASA ORIZON", "CASA DO POÇO", "CASA MAR"
]

# Usuários do sistema
USUARIOS = {
    'Legalização': {'senha': 'legalização2025', 'setor': 'legalizacao'},
    'obra': {'senha': 'Obra2025', 'setor': 'obra'},
    'Projetos': {'senha': 'Projetos2025', 'setor': 'projeto'},
    'marketing': {'senha': 'Marketing2025', 'setor': 'marketing'},
    'Produtos': {'senha': 'Produtos2025', 'setor': 'Marketing'},
    'Pos obra': {'senha': 'Pos-obra2025', 'setor': 'Pos obra'},
    'ADM': {'senha': 'C2025asaOrange', 'setor': 'Adiministrador'}
}

# Gerar matriz de meses até 12/2030
def gerar_meses():
    meses = []
    anos = range(2025, 2031)  # 2025 até 2030
    nomes_meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 
                   'jul', 'ago', 'set', 'out', 'nov', 'dez']
    
    for ano in anos:
        for i, mes_nome in enumerate(nomes_meses, 1):
            if ano == 2025 and i < 11:  # Começa em nov/2025
                continue
            if ano == 2030 and i > 12:  # Até dez/2030
                break
            meses.append(f"{mes_nome}/{str(ano)[2:]}")
    
    return meses

# Template HTML
LOGIN_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Login - Projetado INCORPORAÇÃO</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }
        .login-container {
            background: white;
            border-radius: 15px;
            box-shadow: 0 20px 40px rgba(0, 0, 0, 0.1);
            width: 100%;
            max-width: 400px;
            padding: 40px;
        }
        .logo { text-align: center; margin-bottom: 30px; }
        .logo h1 { color: #E65100; font-size: 32px; margin-bottom: 10px; }
        .logo p { color: #6b7280; font-size: 14px; }
        .form-group { margin-bottom: 20px; }
        .form-group label { display: block; margin-bottom: 8px; color: #374151; font-weight: 500; }
        .form-group input {
            width: 100%; padding: 12px 15px; border: 1px solid #d1d5db;
            border-radius: 8px; font-size: 16px; transition: all 0.3s ease;
        }
        .form-group input:focus {
            outline: none; border-color: #E65100; box-shadow: 0 0 0 3px rgba(230, 81, 0, 0.1);
        }
        .btn-login {
            width: 100%; padding: 14px; background: linear-gradient(135deg, #E65100 0%, #FF9800 100%);
            color: white; border: none; border-radius: 8px; font-size: 16px;
            font-weight: 600; cursor: pointer; transition: all 0.3s ease;
        }
        .btn-login:hover { transform: translateY(-2px); box-shadow: 0 10px 20px rgba(230, 81, 0, 0.2); }
        .error-message {
            background-color: #fee2e2; color: #991b1b; padding: 12px;
            border-radius: 8px; margin-bottom: 20px; text-align: center; display: none;
        }
        .footer { text-align: center; margin-top: 30px; color: #6b7280; font-size: 14px; }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="logo">
            <h1>Projetado INCORPORAÇÃO</h1>
            <p>Sistema de Fluxo de Caixa - Controle Financeiro</p>
        </div>
        <div id="error-message" class="error-message"></div>
        <form id="login-form">
            <div class="form-group">
                <label for="usuario">Usuário</label>
                <input type="text" id="usuario" name="usuario" required placeholder="Digite seu usuário">
            </div>
            <div class="form-group">
                <label for="senha">Senha</label>
                <input type="password" id="senha" name="senha" required placeholder="Digite sua senha">
            </div>
            <button type="submit" class="btn-login">Entrar no Sistema</button>
        </form>
        <div class="footer">
            <p>Use suas credenciais do sistema principal</p>
        </div>
    </div>
    <script>
        document.getElementById('login-form').addEventListener('submit', async function(e) {
            e.preventDefault();
            const usuario = document.getElementById('usuario').value;
            const senha = document.getElementById('senha').value;
            const errorDiv = document.getElementById('error-message');
            errorDiv.style.display = 'none';
            try {
                const response = await fetch('/auth', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ usuario, senha })
                });
                const data = await response.json();
                if (data.success) {
                    window.location.href = data.redirect_url;
                } else {
                    errorDiv.textContent = data.message;
                    errorDiv.style.display = 'block';
                }
            } catch (error) {
                errorDiv.textContent = 'Erro de conexão com o servidor';
                errorDiv.style.display = 'block';
            }
        });
    </script>
</body>
</html>
'''

SISTEMA_TEMPLATE = '''
<!DOCTYPE html>
<html lang="pt-br">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Projetado INCORPORAÇÃO - Sistema de Fluxo de Caixa</title>
    <style>
        :root {
            --primary-color: #E65100;
            --primary-dark: #BF360C;
            --primary-light: #FF9800;
            --secondary-color: #1976D2;
            --success-color: #4CAF50;
            --warning-color: #FFC107;
            --danger-color: #F44336;
            --light-color: #F5F5F5;
            --dark-color: #212121;
            --gray-color: #757575;
            --border-color: #E0E0E0;
            --shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
        }
        
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Roboto', sans-serif;
            background: var(--light-color);
            min-height: 100vh;
            color: var(--dark-color);
        }
        
        /* Header */
        .app-header {
            background: var(--primary-color);
            color: white;
            padding: 1rem 2rem;
            box-shadow: var(--shadow);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .header-content h1 {
            font-size: 1.8rem;
            font-weight: 700;
            margin-bottom: 0.3rem;
        }
        
        .header-content .subtitle {
            font-size: 0.9rem;
            opacity: 0.9;
        }
        
        .auth-section {
            display: flex;
            gap: 10px;
            align-items: center;
        }
        
        .admin-info {
            display: flex;
            align-items: center;
            gap: 10px;
            color: white;
            font-weight: 600;
        }
        
        .admin-badge {
            background-color: var(--success-color);
            padding: 4px 8px;
            border-radius: 20px;
            font-size: 11px;
        }
        
        .user-badge {
            background-color: var(--secondary-color);
            padding: 4px 8px;
            border-radius: 20px;
            font-size: 11px;
        }
        
        .btn {
            padding: 8px 16px;
            border: none;
            border-radius: 4px;
            font-weight: 500;
            cursor: pointer;
            transition: all 0.3s ease;
            display: flex;
            align-items: center;
            gap: 5px;
            font-size: 13px;
        }
        
        .btn-logout {
            background: var(--danger-color);
            color: white;
        }
        
        .btn-logout:hover {
            background: #d32f2f;
        }
        
        /* Container Principal */
        .container {
            max-width: 95%;
            margin: 1rem auto;
            padding: 0 0.5rem;
        }
        
        /* Painel de Controle */
        .control-panel {
            background: white;
            border-radius: 8px;
            padding: 0.8rem;
            margin-bottom: 1rem;
            box-shadow: var(--shadow);
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 0.8rem;
        }
        
        .controls {
            display: flex;
            gap: 0.5rem;
            flex-wrap: wrap;
        }
        
        .btn-primary { background: var(--primary-color); color: white; }
        .btn-success { background: var(--success-color); color: white; }
        .btn-warning { background: var(--warning-color); color: var(--dark-color); }
        .btn-danger { background: var(--danger-color); color: white; }
        .btn-secondary { background: var(--secondary-color); color: white; }
        .btn-info { background: #0ea5e9; color: white; }
        
        .btn:hover {
            transform: translateY(-1px);
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }
        
        /* Status Message */
        .status-message {
            background: #E3F2FD;
            border-left: 4px solid var(--secondary-color);
            padding: 0.6rem;
            margin-bottom: 0.8rem;
            border-radius: 4px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            animation: slideIn 0.3s ease;
            font-size: 0.9rem;
        }
        
        .status-message.success {
            background: #E8F5E9;
            border-left-color: var(--success-color);
        }
        
        .status-message.error {
            background: #FFEBEE;
            border-left-color: var(--danger-color);
        }
        
        .close-btn {
            background: none;
            border: none;
            font-size: 1.2rem;
            cursor: pointer;
            color: var(--gray-color);
        }
        
        .hidden { display: none !important; }
        
        /* Tabela de Controle */
        .grid-container {
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: var(--shadow);
            margin-bottom: 1rem;
        }
        
        .table-wrapper {
            overflow-x: auto;
            max-height: 70vh;
            overflow-y: auto;
            position: relative;
        }
        
        #fluxoTable {
            width: 100%;
            border-collapse: collapse;
            min-width: 1800px;
            font-size: 12px;
        }
        
        #fluxoTable th {
            background: var(--dark-color);
            color: white;
            padding: 0.6rem;
            text-align: center;
            font-weight: 500;
            position: sticky;
            top: 0;
            z-index: 10;
            min-width: 100px;
            border: 1px solid var(--border-color);
        }
        
        #fluxoTable th.date-column {
            background: var(--primary-color);
            position: sticky;
            left: 0;
            z-index: 20;
            min-width: 80px;
            font-weight: 600;
        }
        
        #fluxoTable th.saldo-column {
            background: #2E7D32;
        }
        
        #fluxoTable th.entrada-column {
            background: #1976D2;
        }
        
        #fluxoTable td {
            padding: 0.4rem;
            text-align: center;
            border: 1px solid var(--border-color);
            font-size: 11px;
        }
        
        #fluxoTable tbody tr:nth-child(even) {
            background-color: #F9F9F9;
        }
        
        #fluxoTable tbody tr:hover {
            background-color: #F0F0F0;
        }
        
        #fluxoTable td:first-child {
            background: var(--light-color);
            font-weight: 500;
            position: sticky;
            left: 0;
            z-index: 5;
            min-width: 80px;
        }
        
        /* Células de entrada */
        .input-cell {
            width: 100%;
            padding: 0.3rem;
            border: 1px solid var(--border-color);
            border-radius: 3px;
            text-align: center;
            font-size: 11px;
            background: white;
            min-height: 28px;
        }
        
        .input-cell:focus {
            outline: none;
            border-color: var(--primary-color);
            box-shadow: 0 0 0 2px rgba(230, 81, 0, 0.1);
        }
        
        .input-cell:disabled {
            background: #f8f9fa;
            color: #6c757d;
            cursor: not-allowed;
        }
        
        /* Células de saldo */
        .saldo-cell {
            font-weight: 600;
            padding: 0.4rem;
            border-radius: 3px;
            background: #F1F8E9;
            color: #2E7D32;
            min-height: 28px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        .saldo-negativo {
            background: #FFEBEE;
            color: #C62828;
        }
        
        .saldo-positivo {
            background: #E8F5E9;
            color: #2E7D32;
        }
        
        /* Status Bar */
        .status-bar {
            background: var(--dark-color);
            color: white;
            padding: 0.8rem;
            display: flex;
            justify-content: space-around;
            flex-wrap: wrap;
            gap: 0.8rem;
            border-radius: 0 0 8px 8px;
            font-size: 0.8rem;
        }
        
        .stat-item {
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        
        .stat-value {
            font-size: 1rem;
            font-weight: 700;
            color: var(--primary-light);
        }
        
        .stat-label {
            font-size: 0.7rem;
            opacity: 0.8;
        }
        
        /* Modal de Saldo */
        .modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.5);
            z-index: 1000;
            justify-content: center;
            align-items: center;
        }
        
        .modal-content {
            background: white;
            border-radius: 8px;
            padding: 1.5rem;
            width: 90%;
            max-width: 400px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.2);
        }
        
        .modal-title {
            font-size: 1.2rem;
            margin-bottom: 1rem;
            color: var(--dark-color);
        }
        
        .modal-input {
            width: 100%;
            padding: 0.6rem;
            margin: 0.8rem 0;
            border: 1px solid var(--border-color);
            border-radius: 4px;
            font-size: 0.9rem;
        }
        
        .modal-buttons {
            display: flex;
            justify-content: flex-end;
            gap: 0.8rem;
            margin-top: 1rem;
        }
        
        /* Modal Estatísticas */
        #modalEstatisticas .modal-content {
            max-width: 600px;
        }
        
        #estatisticasContent {
            max-height: 400px;
            overflow-y: auto;
        }
        
        /* Loading */
        .loading-overlay {
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.7);
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            z-index: 1000;
            display: none;
        }
        
        .loading-spinner {
            width: 30px;
            height: 30px;
            border: 3px solid var(--light-color);
            border-top: 3px solid var(--primary-color);
            border-radius: 50%;
            animation: spin 1s linear infinite;
        }
        
        /* Animações */
        @keyframes slideIn {
            from { opacity: 0; transform: translateY(-10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        /* Responsivo */
        @media (max-width: 768px) {
            .control-panel {
                flex-direction: column;
                align-items: stretch;
            }
            
            .controls {
                justify-content: center;
            }
            
            .table-wrapper {
                max-height: 50vh;
            }
            
            .app-header {
                flex-direction: column;
                gap: 0.8rem;
                text-align: center;
            }
            
            .container {
                max-width: 100%;
                padding: 0 0.2rem;
            }
        }
    </style>
</head>
<body>
    <!-- Header -->
    <header class="app-header">
        <div class="header-content">
            <h1>Projetado - INCORPORAÇÃO</h1>
            <p class="subtitle">Sistema de Fluxo de Caixa - Controle Financeiro</p>
        </div>
        <div class="auth-section">
            {% if is_admin %}
            <div class="admin-info">
                <span>Administrador</span>
                <div class="admin-badge">ADM</div>
            </div>
            {% else %}
            <div class="admin-info">
                <span>{{ usuario }}</span>
                <div class="user-badge">{{ setor }}</div>
            </div>
            {% endif %}
            <button class="btn btn-logout" onclick="logout()">
                Sair
            </button>
        </div>
    </header>
    
    <!-- Container Principal -->
    <main class="container">
        <!-- Status Message -->
        <div id="statusMessage" class="status-message hidden">
            <span id="statusText"></span>
            <button onclick="hideStatus()" class="close-btn">&times;</button>
        </div>
        
        <!-- Painel de Controle -->
        <div class="control-panel">
            <div class="controls">
                <button class="btn btn-secondary" onclick="carregarTemplate()">
                    📋 Carregar Template
                </button>
                <button class="btn btn-warning" onclick="validarDados()">
                    ✅ Validar Dados
                </button>
                {% if is_admin %}
                <button class="btn btn-primary" onclick="adicionarSaldo()">
                    💰 Adicionar Saldo
                </button>
                <button class="btn btn-danger" onclick="limparTudo()">
                    🗑️ Limpar Tudo
                </button>
                {% endif %}
            </div>
            <div class="controls">
                <button class="btn btn-success" onclick="exportarExcelCompleto()">
                    📊 Gerar Excel
                </button>
                <button class="btn btn-info" onclick="mostrarEstatisticas()">
                    📈 Estatísticas
                </button>
            </div>
        </div>
        
        <!-- Tabela de Controle -->
        <div class="grid-container">
            <div class="table-wrapper">
                <table id="fluxoTable">
                    <thead id="tableHead"></thead>
                    <tbody id="tableBody"></tbody>
                </table>
            </div>
        </div>
        
        <!-- Status Bar -->
        <div class="status-bar">
            <div class="stat-item">
                <span class="stat-label">Total Empreendimentos</span>
                <span class="stat-value" id="totalEmpreendimentos">8</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">Meses</span>
                <span class="stat-value" id="totalMeses">0</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">Total Entradas</span>
                <span class="stat-value" id="totalEntradas">0</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">Saldo Total</span>
                <span class="stat-value" id="saldoTotal">0</span>
            </div>
            <div class="stat-item">
                <span class="stat-label">API Status</span>
                <span class="stat-value status-online" id="apiStatus">Online</span>
            </div>
        </div>
    </main>
    
    <!-- Modal Adicionar Saldo -->
    <div id="modalSaldo" class="modal">
        <div class="modal-content">
            <h3 class="modal-title">Adicionar Saldo Inicial</h3>
            <div id="saldoInputs"></div>
            <div class="modal-buttons">
                <button class="btn btn-secondary" onclick="fecharModalSaldo()">Cancelar</button>
                <button class="btn btn-success" onclick="salvarSaldosIniciais()">Salvar</button>
            </div>
        </div>
    </div>
    
    <!-- Modal Estatísticas -->
    <div id="modalEstatisticas" class="modal">
        <div class="modal-content">
            <h3 class="modal-title">Estatísticas do Sistema</h3>
            <div id="estatisticasContent"></div>
            <div class="modal-buttons">
                <button class="btn btn-secondary" onclick="fecharModalEstatisticas()">Fechar</button>
            </div>
        </div>
    </div>
    
    <!-- Loading Overlay -->
    <div id="loadingOverlay" class="loading-overlay">
        <div class="loading-spinner"></div>
        <p style="color: white; margin-top: 0.8rem; font-size: 0.9rem;">Processando...</p>
    </div>
    
    <script>
        // Configurações
        let dados = [];
        let meses = [];
        let empreendimentos = {{ empreendimentos|tojson }};
        let isAdmin = {{ 'true' if is_admin else 'false' }};
        let saldosIniciais = {{ saldos_iniciais|tojson }};
        
        // Funções de UI
        function showLoading(show) {
            document.getElementById('loadingOverlay').style.display = show ? 'flex' : 'none';
        }
        
        function showStatus(message, type = 'success') {
            const statusDiv = document.getElementById('statusMessage');
            const statusText = document.getElementById('statusText');
            
            statusText.textContent = message;
            statusDiv.className = `status-message ${type}`;
            statusDiv.classList.remove('hidden');
            
            setTimeout(hideStatus, 5000);
        }
        
        function hideStatus() {
            document.getElementById('statusMessage').classList.add('hidden');
        }
        
        // Carregar dados
        async function carregarDados() {
            showLoading(true);
            try {
                const response = await fetch('/api/dados');
                const result = await response.json();
                
                if (result.success) {
                    dados = result.dados || [];
                    meses = result.meses || [];
                    saldosIniciais = result.saldos_iniciais || {};
                    renderizarTabela();
                    atualizarEstatisticas();
                }
            } catch (error) {
                console.error('Erro ao carregar dados:', error);
                showStatus('Erro ao carregar dados', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Renderizar tabela com 2 colunas por empreendimento
        function renderizarTabela() {
            const tableHead = document.getElementById('tableHead');
            const tableBody = document.getElementById('tableBody');
            
            // Limpar tabela
            tableHead.innerHTML = '';
            tableBody.innerHTML = '';
            
            // Criar cabeçalho
            let headerRow = document.createElement('tr');
            
            // Coluna de datas
            let thDate = document.createElement('th');
            thDate.className = 'date-column';
            thDate.textContent = 'Datas';
            headerRow.appendChild(thDate);
            
            // Colunas para cada empreendimento (Saldo + Entrada)
            empreendimentos.forEach(empreendimento => {
                // Coluna do Saldo
                let thSaldo = document.createElement('th');
                thSaldo.className = 'saldo-column';
                thSaldo.textContent = `${empreendimento} - Saldo`;
                thSaldo.title = 'Saldo disponível (somente admin pode editar)';
                headerRow.appendChild(thSaldo);
                
                // Coluna da Entrada
                let thEntrada = document.createElement('th');
                thEntrada.className = 'entrada-column';
                thEntrada.textContent = `${empreendimento} - Entrada`;
                thEntrada.title = 'Valor a ser subtraído do saldo (todos podem editar)';
                headerRow.appendChild(thEntrada);
            });
            
            tableHead.appendChild(headerRow);
            
            // Linhas de meses
            meses.forEach((mes, mesIndex) => {
                let row = document.createElement('tr');
                
                // Coluna da data
                let tdDate = document.createElement('td');
                tdDate.textContent = mes;
                tdDate.style.fontWeight = 'bold';
                row.appendChild(tdDate);
                
                // Colunas para cada empreendimento
                empreendimentos.forEach(empreendimento => {
                    // Encontrar dados deste mês/empreendimento
                    let dadosMes = dados.find(d => d.mes === mes && d.empreendimento === empreendimento);
                    let entradaAtual = dadosMes?.entrada || 0;
                    
                    // Calcular saldo acumulado até este mês
                    let saldoAcumulado = calcularSaldoAcumulado(mes, empreendimento);
                    
                    // Saldo Atual (editável apenas por admin)
                    let tdSaldo = document.createElement('td');
                    if (isAdmin) {
                        let inputSaldo = document.createElement('input');
                        inputSaldo.type = 'number';
                        inputSaldo.className = 'input-cell';
                        inputSaldo.value = saldoAcumulado;
                        inputSaldo.step = '0.01';
                        inputSaldo.min = '0';
                        inputSaldo.dataset.mes = mes;
                        inputSaldo.dataset.empreendimento = empreendimento;
                        inputSaldo.dataset.tipo = 'saldo';
                        inputSaldo.addEventListener('change', function(e) {
                            atualizarSaldo(mes, empreendimento, parseFloat(e.target.value) || 0);
                        });
                        tdSaldo.appendChild(inputSaldo);
                    } else {
                        tdSaldo.className = `saldo-cell ${saldoAcumulado < 0 ? 'saldo-negativo' : 'saldo-positivo'}`;
                        tdSaldo.textContent = formatarNumero(saldoAcumulado);
                    }
                    row.appendChild(tdSaldo);
                    
                    // Entrada (input) - TODOS podem editar
                    let tdEntrada = document.createElement('td');
                    let inputEntrada = document.createElement('input');
                    inputEntrada.type = 'number';
                    inputEntrada.className = 'input-cell';
                    inputEntrada.value = entradaAtual;
                    inputEntrada.step = '0.01';
                    inputEntrada.min = '0';
                    inputEntrada.placeholder = '0.00';
                    inputEntrada.dataset.mes = mes;
                    inputEntrada.dataset.empreendimento = empreendimento;
                    inputEntrada.dataset.tipo = 'entrada';
                    
                    // Evento para atualizar entrada
                    inputEntrada.addEventListener('change', function(e) {
                        atualizarEntrada(mes, empreendimento, parseFloat(e.target.value) || 0);
                    });
                    
                    tdEntrada.appendChild(inputEntrada);
                    row.appendChild(tdEntrada);
                });
                
                tableBody.appendChild(row);
            });
        }
        
        // Calcular saldo acumulado até um mês específico
        function calcularSaldoAcumulado(mes, empreendimento) {
            // Saldo inicial
            let saldo = saldosIniciais[empreendimento] || 0;
            
            // Encontrar índice do mês atual
            const mesIndex = meses.indexOf(mes);
            if (mesIndex === -1) return saldo;
            
            // Subtrair todas as entradas até este mês
            for (let i = 0; i <= mesIndex; i++) {
                const mesAtual = meses[i];
                const dadosMes = dados.find(d => d.mes === mesAtual && d.empreendimento === empreendimento);
                if (dadosMes && dadosMes.entrada) {
                    saldo -= dadosMes.entrada;
                }
            }
            
            return saldo;
        }
        
        // Atualizar entrada (TODOS podem fazer)
        async function atualizarEntrada(mes, empreendimento, valor) {
            if (valor < 0) {
                showStatus('Valor de entrada não pode ser negativo', 'error');
                return;
            }
            
            showLoading(true);
            
            try {
                const response = await fetch('/api/atualizar-entrada', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        mes: mes,
                        empreendimento: empreendimento,
                        entrada: valor
                    })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    // Atualizar dados locais
                    let index = dados.findIndex(d => d.mes === mes && d.empreendimento === empreendimento);
                    
                    if (index >= 0) {
                        dados[index].entrada = valor;
                    } else {
                        dados.push({
                            mes: mes,
                            empreendimento: empreendimento,
                            entrada: valor
                        });
                    }
                    
                    // Recalcular e renderizar tabela
                    renderizarTabela();
                    atualizarEstatisticas();
                    showStatus('Entrada atualizada com sucesso');
                } else {
                    showStatus('Erro ao atualizar: ' + result.error, 'error');
                }
            } catch (error) {
                console.error('Erro ao atualizar entrada:', error);
                showStatus('Erro ao atualizar entrada', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Atualizar saldo (APENAS ADMIN)
        async function atualizarSaldo(mes, empreendimento, valor) {
            if (!isAdmin) {
                showStatus('Apenas administradores podem atualizar saldos', 'error');
                return;
            }
            
            showLoading(true);
            
            try {
                const response = await fetch('/api/atualizar-saldo', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        empreendimento: empreendimento,
                        saldo: valor
                    })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    saldosIniciais[empreendimento] = valor;
                    renderizarTabela();
                    atualizarEstatisticas();
                    showStatus('Saldo inicial atualizado');
                }
            } catch (error) {
                console.error('Erro ao atualizar saldo:', error);
                showStatus('Erro ao atualizar saldo', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Carregar template
        async function carregarTemplate() {
            showLoading(true);
            
            try {
                const response = await fetch('/api/carregar-template', { method: 'POST' });
                const result = await response.json();
                
                if (result.success) {
                    meses = result.meses;
                    saldosIniciais = result.saldos_iniciais || {};
                    dados = [];
                    renderizarTabela();
                    atualizarEstatisticas();
                    showStatus('Template carregado com sucesso');
                }
            } catch (error) {
                console.error('Erro ao carregar template:', error);
                showStatus('Erro ao carregar template', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Validar dados
        async function validarDados() {
            showLoading(true);
            
            try {
                const response = await fetch('/api/validar-dados', { method: 'POST' });
                const result = await response.json();
                
                if (result.success) {
                    if (result.erros && result.erros.length > 0) {
                        showStatus('Dados validados com ' + result.erros.length + ' erros', 'error');
                        console.warn('Erros encontrados:', result.erros);
                    } else {
                        showStatus('Todos os dados estão válidos');
                    }
                }
            } catch (error) {
                console.error('Erro ao validar dados:', error);
                showStatus('Erro ao validar dados', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Adicionar saldo (modal)
        function adicionarSaldo() {
            if (!isAdmin) {
                showStatus('Apenas administradores podem adicionar saldo', 'error');
                return;
            }
            
            const modal = document.getElementById('modalSaldo');
            const inputsDiv = document.getElementById('saldoInputs');
            inputsDiv.innerHTML = '';
            
            empreendimentos.forEach(empreendimento => {
                const div = document.createElement('div');
                div.style.marginBottom = '8px';
                div.style.display = 'flex';
                div.style.alignItems = 'center';
                
                const label = document.createElement('label');
                label.textContent = `${empreendimento}: `;
                label.style.display = 'inline-block';
                label.style.width = '120px';
                label.style.fontSize = '0.9rem';
                
                const input = document.createElement('input');
                input.type = 'number';
                input.className = 'modal-input';
                input.value = saldosIniciais[empreendimento] || 0;
                input.step = '0.01';
                input.min = '0';
                input.dataset.empreendimento = empreendimento;
                input.style.flex = '1';
                
                div.appendChild(label);
                div.appendChild(input);
                inputsDiv.appendChild(div);
            });
            
            modal.style.display = 'flex';
        }
        
        function fecharModalSaldo() {
            document.getElementById('modalSaldo').style.display = 'none';
        }
        
        async function salvarSaldosIniciais() {
            showLoading(true);
            
            try {
                const inputs = document.querySelectorAll('#saldoInputs input');
                const novosSaldos = {};
                
                inputs.forEach(input => {
                    const empreendimento = input.dataset.empreendimento;
                    novosSaldos[empreendimento] = parseFloat(input.value) || 0;
                });
                
                const response = await fetch('/api/atualizar-saldos', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ saldos: novosSaldos })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    saldosIniciais = result.saldos;
                    fecharModalSaldo();
                    renderizarTabela();
                    atualizarEstatisticas();
                    showStatus('Saldos iniciais atualizados');
                }
            } catch (error) {
                console.error('Erro ao salvar saldos:', error);
                showStatus('Erro ao salvar saldos', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Limpar tudo (apenas admin)
        async function limparTudo() {
            if (!isAdmin) {
                showStatus('Apenas administradores podem limpar dados', 'error');
                return;
            }
            
            if (!confirm('Tem certeza que deseja limpar todos os dados? Esta ação não pode ser desfeita.')) {
                return;
            }
            
            showLoading(true);
            
            try {
                const response = await fetch('/api/limpar-tudo', { method: 'POST' });
                const result = await response.json();
                
                if (result.success) {
                    dados = [];
                    saldosIniciais = result.saldos_iniciais || {};
                    renderizarTabela();
                    atualizarEstatisticas();
                    showStatus('Todos os dados foram limpos');
                }
            } catch (error) {
                console.error('Erro ao limpar dados:', error);
                showStatus('Erro ao limpar dados', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Exportar Excel com estrutura completa
        async function exportarExcelCompleto() {
            showLoading(true);
            
            try {
                const response = await fetch('/api/exportar-excel-completo');
                const blob = await response.blob();
                
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `Fluxo_Caixa_Projetado_${new Date().toISOString().split('T')[0]}.xlsx`;
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                
                showStatus('Excel exportado com estrutura completa!');
            } catch (error) {
                console.error('Erro ao exportar:', error);
                showStatus('Erro ao exportar Excel', 'error');
            } finally {
                showLoading(false);
            }
        }
        
        // Mostrar estatísticas
        async function mostrarEstatisticas() {
            const modal = document.getElementById('modalEstatisticas');
            const contentDiv = document.getElementById('estatisticasContent');
            
            // Calcular estatísticas
            let totalEntradas = 0;
            let saldoTotal = 0;
            let empreendimentosStats = {};
            
            empreendimentos.forEach(empreendimento => {
                const saldoInicial = saldosIniciais[empreendimento] || 0;
                let entradasEmpreendimento = 0;
                
                dados.filter(d => d.empreendimento === empreendimento).forEach(d => {
                    entradasEmpreendimento += d.entrada || 0;
                });
                
                totalEntradas += entradasEmpreendimento;
                const saldoFinal = saldoInicial - entradasEmpreendimento;
                saldoTotal += saldoFinal;
                
                empreendimentosStats[empreendimento] = {
                    saldo_inicial: saldoInicial,
                    entradas: entradasEmpreendimento,
                    saldo_final: saldoFinal
                };
            });
            
            // Construir conteúdo
            let html = `
                <div style="margin-bottom: 1rem;">
                    <h4 style="color: var(--primary-color); margin-bottom: 0.5rem;">Resumo Geral</h4>
                    <p><strong>Total de Empreendimentos:</strong> ${empreendimentos.length}</p>
                    <p><strong>Total de Meses:</strong> ${meses.length}</p>
                    <p><strong>Total de Entradas:</strong> ${formatarNumero(totalEntradas)}</p>
                    <p><strong>Saldo Total:</strong> <span style="color: ${saldoTotal >= 0 ? '#2E7D32' : '#C62828'}">${formatarNumero(saldoTotal)}</span></p>
                </div>
                
                <div style="margin-bottom: 1rem;">
                    <h4 style="color: var(--primary-color); margin-bottom: 0.5rem;">Por Empreendimento</h4>
            `;
            
            empreendimentos.forEach(empreendimento => {
                const stats = empreendimentosStats[empreendimento];
                html += `
                    <div style="background: #f5f5f5; padding: 0.5rem; margin-bottom: 0.5rem; border-radius: 4px;">
                        <strong>${empreendimento}</strong>
                        <div style="display: flex; justify-content: space-between; font-size: 0.9rem;">
                            <span>Saldo Inicial: ${formatarNumero(stats.saldo_inicial)}</span>
                            <span>Entradas: ${formatarNumero(stats.entradas)}</span>
                            <span style="color: ${stats.saldo_final >= 0 ? '#2E7D32' : '#C62828'}">
                                Saldo Final: ${formatarNumero(stats.saldo_final)}
                            </span>
                        </div>
                    </div>
                `;
            });
            
            html += `</div>`;
            
            contentDiv.innerHTML = html;
            modal.style.display = 'flex';
        }
        
        function fecharModalEstatisticas() {
            document.getElementById('modalEstatisticas').style.display = 'none';
        }
        
        // Atualizar estatísticas
        function atualizarEstatisticas() {
            // Totais
            let totalEntradas = 0;
            let saldoTotal = 0;
            
            dados.forEach(d => {
                totalEntradas += d.entrada || 0;
            });
            
            empreendimentos.forEach(empreendimento => {
                const saldoInicial = saldosIniciais[empreendimento] || 0;
                let entradasEmpreendimento = 0;
                
                dados.filter(d => d.empreendimento === empreendimento).forEach(d => {
                    entradasEmpreendimento += d.entrada || 0;
                });
                
                saldoTotal += (saldoInicial - entradasEmpreendimento);
            });
            
            document.getElementById('totalEmpreendimentos').textContent = empreendimentos.length;
            document.getElementById('totalMeses').textContent = meses.length;
            document.getElementById('totalEntradas').textContent = formatarNumero(totalEntradas);
            document.getElementById('saldoTotal').textContent = formatarNumero(saldoTotal);
        }
        
        // Utilitários
        function formatarNumero(valor) {
            return new Intl.NumberFormat('pt-BR', {
                minimumFractionDigits: 2,
                maximumFractionDigits: 2
            }).format(valor);
        }
        
        function logout() {
            window.location.href = '/logout';
        }
        
        // Inicialização
        document.addEventListener('DOMContentLoaded', () => {
            carregarDados();
            
            // Atualizar estatísticas periodicamente
            setInterval(atualizarEstatisticas, 30000);
        });
    </script>
</body>
</html>
'''

# Funções auxiliares
def carregar_dados():
    """Carrega dados do arquivo JSON"""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('dados', []), data.get('meses', gerar_meses()), data.get('saldos_iniciais', {})
        except:
            return [], gerar_meses(), {}
    return [], gerar_meses(), {}

def salvar_dados(dados, meses, saldos_iniciais):
    """Salva dados no arquivo JSON"""
    data = {
        'dados': dados,
        'meses': meses,
        'saldos_iniciais': saldos_iniciais,
        'ultima_atualizacao': datetime.now().isoformat(),
        'usuario_ultima_alteracao': session.get('usuario', 'desconhecido')
    }
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def calcular_saldo_acumulado_export(mes, empreendimento, dados, meses, saldos_iniciais):
    """Calcula saldo acumulado para exportação"""
    saldo = saldos_iniciais.get(empreendimento, 0)
    mes_index = meses.index(mes)
    
    for i in range(mes_index + 1):
        mes_atual = meses[i]
        dados_mes = next((d for d in dados if d['mes'] == mes_atual and d['empreendimento'] == empreendimento), None)
        if dados_mes:
            saldo -= dados_mes.get('entrada', 0)
    
    return saldo

# Decoradores
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login_page'))
        if session.get('usuario') != 'ADM':
            return jsonify({'success': False, 'error': 'Acesso restrito a administradores'}), 403
        return f(*args, **kwargs)
    return decorated_function

# Rotas
@app.route('/')
@app.route('/login')
def login_page():
    if 'usuario' in session:
        return redirect(url_for('sistema_controle'))
    return render_template_string(LOGIN_TEMPLATE)

@app.route('/auth', methods=['POST'])
def auth():
    data = request.get_json()
    usuario = data.get('usuario', '')
    senha = data.get('senha', '')
    
    if usuario in USUARIOS and USUARIOS[usuario]['senha'] == senha:
        session['usuario'] = usuario
        session['setor'] = USUARIOS[usuario]['setor']
        session['is_admin'] = (usuario == 'ADM')
        
        return jsonify({
            'success': True,
            'redirect_url': '/sistema-controle'
        })
    
    return jsonify({
        'success': False,
        'message': 'Usuário ou senha incorretos'
    }), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/sistema-controle')
@login_required
def sistema_controle():
    dados, meses, saldos_iniciais = carregar_dados()
    
    # Se não tiver meses, carrega template
    if not meses:
        meses = gerar_meses()
    
    # Se não tiver saldos iniciais, inicializa com zeros
    if not saldos_iniciais:
        saldos_iniciais = {emp: 0 for emp in EMPREENDIMENTOS}
    
    return render_template_string(SISTEMA_TEMPLATE, 
                                usuario=session['usuario'],
                                setor=session['setor'],
                                is_admin=session.get('is_admin', False),
                                empreendimentos=EMPREENDIMENTOS,
                                saldos_iniciais=saldos_iniciais)

@app.route('/api/dados')
@login_required
def api_dados():
    dados, meses, saldos_iniciais = carregar_dados()
    return jsonify({
        'success': True,
        'dados': dados,
        'meses': meses,
        'saldos_iniciais': saldos_iniciais
    })

@app.route('/api/atualizar-entrada', methods=['POST'])
@login_required
def api_atualizar_entrada():
    """Atualiza entrada - TODOS os usuários podem fazer"""
    try:
        data = request.get_json()
        mes = data.get('mes')
        empreendimento = data.get('empreendimento')
        entrada = float(data.get('entrada', 0))
        
        if not mes or not empreendimento:
            return jsonify({'success': False, 'error': 'Dados incompletos'}), 400
        
        if entrada < 0:
            return jsonify({'success': False, 'error': 'Entrada não pode ser negativa'}), 400
        
        dados, meses, saldos_iniciais = carregar_dados()
        
        # Verificar se já existe registro para este mês/empreendimento
        index = next((i for i, d in enumerate(dados) 
                     if d['mes'] == mes and d['empreendimento'] == empreendimento), -1)
        
        if index >= 0:
            dados[index]['entrada'] = entrada
            dados[index]['usuario'] = session.get('usuario')
            dados[index]['data_atualizacao'] = datetime.now().isoformat()
        else:
            dados.append({
                'mes': mes,
                'empreendimento': empreendimento,
                'entrada': entrada,
                'usuario': session.get('usuario'),
                'data_atualizacao': datetime.now().isoformat()
            })
        
        salvar_dados(dados, meses, saldos_iniciais)
        
        return jsonify({
            'success': True,
            'message': 'Entrada atualizada com sucesso'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/atualizar-saldo', methods=['POST'])
@login_required
@admin_required
def api_atualizar_saldo():
    """Atualiza saldo inicial - APENAS ADMIN pode fazer"""
    try:
        data = request.get_json()
        empreendimento = data.get('empreendimento')
        saldo = float(data.get('saldo', 0))
        
        if not empreendimento:
            return jsonify({'success': False, 'error': 'Empreendimento não informado'}), 400
        
        dados, meses, saldos_iniciais = carregar_dados()
        
        saldos_iniciais[empreendimento] = saldo
        salvar_dados(dados, meses, saldos_iniciais)
        
        return jsonify({
            'success': True,
            'saldos': saldos_iniciais
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/carregar-template', methods=['POST'])
@login_required
def api_carregar_template():
    """Carrega template com matriz completa até 12/2030"""
    try:
        meses = gerar_meses()
        saldos_iniciais = {emp: 0 for emp in EMPREENDIMENTOS}
        
        salvar_dados([], meses, saldos_iniciais)
        
        return jsonify({
            'success': True,
            'meses': meses,
            'saldos_iniciais': saldos_iniciais,
            'message': 'Template carregado com sucesso'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/validar-dados', methods=['POST'])
@login_required
def api_validar_dados():
    """Valida os dados inseridos"""
    try:
        dados, meses, saldos_iniciais = carregar_dados()
        erros = []
        
        # Validar se há entradas negativas
        for d in dados:
            if d.get('entrada', 0) < 0:
                erros.append(f"Entrada negativa em {d['mes']} - {d['empreendimento']}")
        
        # Validar se há saldos iniciais negativos
        for emp, saldo in saldos_iniciais.items():
            if saldo < 0:
                erros.append(f"Saldo inicial negativo para {emp}")
        
        return jsonify({
            'success': True,
            'erros': erros,
            'total_erros': len(erros),
            'message': f'Validação concluída com {len(erros)} erro(s)' if erros else 'Todos os dados estão válidos'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/atualizar-saldos', methods=['POST'])
@login_required
@admin_required
def api_atualizar_saldos():
    """Atualiza todos os saldos iniciais de uma vez"""
    try:
        data = request.get_json()
        novos_saldos = data.get('saldos', {})
        
        dados, meses, saldos_iniciais = carregar_dados()
        
        # Atualizar saldos
        for emp, saldo in novos_saldos.items():
            if emp in EMPREENDIMENTOS:
                saldos_iniciais[emp] = float(saldo)
        
        salvar_dados(dados, meses, saldos_iniciais)
        
        return jsonify({
            'success': True,
            'saldos': saldos_iniciais,
            'message': 'Saldos atualizados com sucesso'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/limpar-tudo', methods=['POST'])
@login_required
@admin_required
def api_limpar_tudo():
    """Limpa todos os dados, mantendo apenas a estrutura"""
    try:
        meses = gerar_meses()
        saldos_iniciais = {emp: 0 for emp in EMPREENDIMENTOS}
        
        salvar_dados([], meses, saldos_iniciais)
        
        return jsonify({
            'success': True,
            'saldos_iniciais': saldos_iniciais,
            'message': 'Todos os dados foram limpos'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exportar-excel-completo')
@login_required
def api_exportar_excel_completo():
    """Exporta dados para Excel mantendo estrutura da tabela web"""
    try:
        dados, meses, saldos_iniciais = carregar_dados()
        
        # Criar workbook em memória
        wb = Workbook()
        ws = wb.active
        ws.title = "Fluxo de Caixa"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        date_font = Font(bold=True, color="000000", size=11)
        saldo_font = Font(bold=True, color="2E7D32", size=11)  # Verde para positivo
        saldo_negativo_font = Font(bold=True, color="C62828", size=11)  # Vermelho para negativo
        entrada_font = Font(color="1976D2", size=11)  # Azul para entradas
        
        # Cores de fundo
        header_fill_saldo = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Verde escuro
        header_fill_entrada = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")  # Azul
        date_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")  # Laranja
        saldo_positivo_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Verde claro
        saldo_negativo_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Vermelho claro
        entrada_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Azul claro
        
        # Bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alinhamento
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        
        # Linha 1: Título
        ws.merge_cells('A1:Q1')
        title_cell = ws['A1']
        title_cell.value = "Projetado - INCORPORAÇÃO - Sistema de Fluxo de Caixa"
        title_cell.font = Font(bold=True, size=14, color="E65100")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Linha 2: Sub-título
        ws.merge_cells('A2:Q2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Usuário: {session.get('usuario', 'N/A')}"
        subtitle_cell.font = Font(italic=True, size=10)
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Linha 4: Cabeçalho das datas
        ws['A4'] = "Datas"
        ws['A4'].font = header_font
        ws['A4'].fill = date_fill
        ws['A4'].border = thin_border
        ws['A4'].alignment = center_alignment
        
        # Cabeçalhos dos empreendimentos (duas colunas cada)
        col_index = 1  # Começa na coluna B (índice 2)
        for i, empreendimento in enumerate(EMPREENDIMENTOS):
            # Coluna do Saldo
            saldo_col = get_column_letter(col_index + 1)
            ws[f'{saldo_col}4'] = f"{empreendimento}\nSaldo"
            ws[f'{saldo_col}4'].font = header_font
            ws[f'{saldo_col}4'].fill = header_fill_saldo
            ws[f'{saldo_col}4'].border = thin_border
            ws[f'{saldo_col}4'].alignment = center_alignment
            ws.column_dimensions[saldo_col].width = 15
            
            # Coluna da Entrada
            entrada_col = get_column_letter(col_index + 2)
            ws[f'{entrada_col}4'] = f"{empreendimento}\nEntrada"
            ws[f'{entrada_col}4'].font = header_font
            ws[f'{entrada_col}4'].fill = header_fill_entrada
            ws[f'{entrada_col}4'].border = thin_border
            ws[f'{entrada_col}4'].alignment = center_alignment
            ws.column_dimensions[entrada_col].width = 15
            
            col_index += 2
        
        # Ajustar largura da coluna A (Datas)
        ws.column_dimensions['A'].width = 12
        
        # Preencher dados por mês
        row_index = 5  # Começa na linha 5
        for mes in meses:
            # Coluna da data
            ws[f'A{row_index}'] = mes
            ws[f'A{row_index}'].font = date_font
            ws[f'A{row_index}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            ws[f'A{row_index}'].border = thin_border
            ws[f'A{row_index}'].alignment = center_alignment
            
            # Preencher dados para cada empreendimento
            col_index = 1  # Reset para coluna B
            for empreendimento in EMPREENDIMENTOS:
                # Calcular saldo acumulado até este mês
                saldo_acumulado = calcular_saldo_acumulado_export(mes, empreendimento, dados, meses, saldos_iniciais)
                
                # Encontrar entrada deste mês
                entrada_atual = 0
                dados_mes = next((d for d in dados if d['mes'] == mes and d['empreendimento'] == empreendimento), None)
                if dados_mes:
                    entrada_atual = dados_mes.get('entrada', 0)
                
                # Coluna do Saldo
                saldo_col = get_column_letter(col_index + 1)
                saldo_cell = ws[f'{saldo_col}{row_index}']
                saldo_cell.value = saldo_acumulado
                saldo_cell.number_format = '#,##0.00'
                
                # Estilo baseado no valor do saldo
                if saldo_acumulado < 0:
                    saldo_cell.font = saldo_negativo_font
                    saldo_cell.fill = saldo_negativo_fill
                else:
                    saldo_cell.font = saldo_font
                    saldo_cell.fill = saldo_positivo_fill
                
                saldo_cell.border = thin_border
                saldo_cell.alignment = right_alignment
                
                # Coluna da Entrada
                entrada_col = get_column_letter(col_index + 2)
                entrada_cell = ws[f'{entrada_col}{row_index}']
                entrada_cell.value = entrada_atual
                entrada_cell.number_format = '#,##0.00'
                entrada_cell.font = entrada_font
                entrada_cell.fill = entrada_fill
                entrada_cell.border = thin_border
                entrada_cell.alignment = right_alignment
                
                col_index += 2
            
            row_index += 1
        
        # Linha de totais
        total_row = row_index + 1
        
        # Célula "TOTAIS"
        ws[f'A{total_row}'] = "TOTAIS"
        ws[f'A{total_row}'].font = Font(bold=True, size=12)
        ws[f'A{total_row}'].fill = PatternFill(start_color="212121", end_color="212121", fill_type="solid")
        ws[f'A{total_row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{total_row}'].border = thin_border
        ws[f'A{total_row}'].alignment = center_alignment
        
        # Calcular totais por empreendimento
        col_index = 1
        for empreendimento in EMPREENDIMENTOS:
            # Total de entradas por empreendimento
            total_entradas = sum(d.get('entrada', 0) for d in dados if d['empreendimento'] == empreendimento)
            
            # Saldo final por empreendimento
            saldo_inicial = saldos_iniciais.get(empreendimento, 0)
            saldo_final = saldo_inicial - total_entradas
            
            # Coluna do Saldo (saldo final)
            saldo_col = get_column_letter(col_index + 1)
            saldo_cell = ws[f'{saldo_col}{total_row}']
            saldo_cell.value = saldo_final
            saldo_cell.number_format = '#,##0.00'
            saldo_cell.font = Font(bold=True, color="2E7D32" if saldo_final >= 0 else "C62828")
            saldo_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            saldo_cell.border = thin_border
            saldo_cell.alignment = right_alignment
            
            # Coluna da Entrada (total de entradas)
            entrada_col = get_column_letter(col_index + 2)
            entrada_cell = ws[f'{entrada_col}{total_row}']
            entrada_cell.value = total_entradas
            entrada_cell.number_format = '#,##0.00'
            entrada_cell.font = Font(bold=True)
            entrada_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            entrada_cell.border = thin_border
            entrada_cell.alignment = right_alignment
            
            col_index += 2
        
        # Adicionar resumo em outra aba
        ws_resumo = wb.create_sheet(title="Resumo")
        
        # Título do resumo
        ws_resumo.merge_cells('A1:D1')
        ws_resumo['A1'] = "Resumo Financeiro por Empreendimento"
        ws_resumo['A1'].font = Font(bold=True, size=14, color="E65100")
        ws_resumo['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeçalhos do resumo
        headers_resumo = ["Empreendimento", "Saldo Inicial", "Total Entradas", "Saldo Final"]
        for i, header in enumerate(headers_resumo, 1):
            cell = ws_resumo.cell(row=3, column=i, value=header)
            cell.font = header_font
            cell.fill = date_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            ws_resumo.column_dimensions[get_column_letter(i)].width = 20
        
        # Dados do resumo
        row_resumo = 4
        for empreendimento in EMPREENDIMENTOS:
            saldo_inicial = saldos_iniciais.get(empreendimento, 0)
            total_entradas = sum(d.get('entrada', 0) for d in dados if d['empreendimento'] == empreendimento)
            saldo_final = saldo_inicial - total_entradas
            
            ws_resumo.cell(row=row_resumo, column=1, value=empreendimento).font = Font(bold=True)
            ws_resumo.cell(row=row_resumo, column=2, value=saldo_inicial).number_format = '#,##0.00'
            ws_resumo.cell(row=row_resumo, column=3, value=total_entradas).number_format = '#,##0.00'
            
            saldo_final_cell = ws_resumo.cell(row=row_resumo, column=4, value=saldo_final)
            saldo_final_cell.number_format = '#,##0.00'
            saldo_final_cell.font = Font(bold=True, color="2E7D32" if saldo_final >= 0 else "C62828")
            
            # Adicionar bordas
            for col in range(1, 5):
                ws_resumo.cell(row=row_resumo, column=col).border = thin_border
            
            row_resumo += 1
        
        # Linha de totais no resumo
        ws_resumo.cell(row=row_resumo, column=1, value="TOTAIS").font = Font(bold=True)
        ws_resumo.cell(row=row_resumo, column=2, value=sum(saldos_iniciais.values())).number_format = '#,##0.00'
        ws_resumo.cell(row=row_resumo, column=3, value=sum(d.get('entrada', 0) for d in dados)).number_format = '#,##0.00'
        
        total_saldo_final = sum(saldos_iniciais.values()) - sum(d.get('entrada', 0) for d in dados)
        total_cell = ws_resumo.cell(row=row_resumo, column=4, value=total_saldo_final)
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(bold=True, color="2E7D32" if total_saldo_final >= 0 else "C62828")
        
        # Ajustar bordas da linha de totais
        for col in range(1, 5):
            ws_resumo.cell(row=row_resumo, column=col).border = thin_border
            ws_resumo.cell(row=row_resumo, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Criar terceira aba com histórico
        ws_historico = wb.create_sheet(title="Histórico")
        
        ws_historico.merge_cells('A1:F1')
        ws_historico['A1'] = "Histórico de Alterações"
        ws_historico['A1'].font = Font(bold=True, size=14, color="E65100")
        ws_historico['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Cabeçalhos do histórico
        headers_historico = ["Data/Hora", "Usuário", "Empreendimento", "Mês", "Entrada", "Tipo"]
        for i, header in enumerate(headers_historico, 1):
            cell = ws_historico.cell(row=3, column=i, value=header)
            cell.font = header_font
            cell.fill = date_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            ws_historico.column_dimensions[get_column_letter(i)].width = 20
        
        # Preencher histórico
        row_hist = 4
        for d in sorted(dados, key=lambda x: x.get('data_atualizacao', ''), reverse=True):
            if d.get('entrada', 0) > 0:  # Só mostrar entradas com valor
                data_str = ''
                if 'data_atualizacao' in d:
                    try:
                        data_obj = datetime.fromisoformat(d['data_atualizacao'].replace('Z', '+00:00'))
                        data_str = data_obj.strftime('%d/%m/%Y %H:%M')
                    except:
                        data_str = d['data_atualizacao']
                
                ws_historico.cell(row=row_hist, column=1, value=data_str)
                ws_historico.cell(row=row_hist, column=2, value=d.get('usuario', ''))
                ws_historico.cell(row=row_hist, column=3, value=d.get('empreendimento', ''))
                ws_historico.cell(row=row_hist, column=4, value=d.get('mes', ''))
                ws_historico.cell(row=row_hist, column=5, value=d.get('entrada', 0)).number_format = '#,##0.00'
                ws_historico.cell(row=row_hist, column=6, value="Entrada")
                
                # Adicionar bordas
                for col in range(1, 7):
                    ws_historico.cell(row=row_hist, column=col).border = thin_border
                
                row_hist += 1
        
        # Salvar para bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Registrar exportação
        print(f"📤 Excel exportado por {session.get('usuario', 'N/A')} em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="Fluxo_Caixa_Projetado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        }
        
    except Exception as e:
        print(f"❌ Erro ao exportar Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 SISTEMA DE FLUXO DE CAIXA - PROJETADO INCORPORAÇÃO")
    print("=" * 60)
    print("📍 Disponível em: http://127.0.0.1:5000")
    print("🔐 Login admin: ADM / C2025asaOrange")
    print("👤 Outros usuários: Podem apenas editar entradas")
    print("📊 Exportação Excel: Mantém estrutura da tabela web")
    print("=" * 60)
    
    # Criar dados iniciais se não existirem
    dados, meses, saldos_iniciais = carregar_dados()
    if not meses:
        meses = gerar_meses()
        saldos_iniciais = {emp: 0 for emp in EMPREENDIMENTOS}
        salvar_dados([], meses, saldos_iniciais)
        print("📁 Template com matriz até 12/2030 criado com sucesso")
    
    print(f"📈 Empreendimentos: {len(EMPREENDIMENTOS)}")
    print(f"📅 Meses gerados: {len(meses)} (nov/2025 até dez/2030)")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=True)