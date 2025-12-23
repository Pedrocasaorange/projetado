import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config import EMPREENDIMENTOS
from utils import calcular_saldo_acumulado_export

def gerar_excel_fluxo(dados, meses, saldos_iniciais, usuario_nome):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fluxo de Caixa"
    
    # --- Configuração de Estilos ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    date_font = Font(bold=True, color="000000", size=11)
    saldo_font = Font(bold=True, color="2E7D32", size=11)
    saldo_negativo_font = Font(bold=True, color="C62828", size=11)
    entrada_font = Font(color="1976D2", size=11)
    
    header_fill_saldo = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_fill_entrada = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    date_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    saldo_positivo_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    saldo_negativo_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    entrada_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # --- Cabeçalho Superior ---
    ws.merge_cells('A1:Q1')
    ws['A1'] = "Projetado - INCORPORAÇÃO - Sistema de Fluxo de Caixa"
    ws['A1'].font = Font(bold=True, size=14, color="E65100")
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Usuário: {usuario_nome}"
    ws['A2'].alignment = center_align

    # --- Títulos das Colunas ---
    ws['A4'] = "Datas"
    ws['A4'].font, ws['A4'].fill, ws['A4'].border, ws['A4'].alignment = header_font, date_fill, thin_border, center_align
    
    col_idx = 1
    for emp in EMPREENDIMENTOS:
        s_col = get_column_letter(col_idx + 1)
        e_col = get_column_letter(col_idx + 2)
        
        ws[f'{s_col}4'] = f"{emp}\nSaldo"
        ws[f'{e_col}4'] = f"{emp}\nEntrada"
        
        for c in [s_col, e_col]:
            ws[f'{c}4'].font = header_font
            ws[f'{c}4'].border = thin_border
            ws[f'{c}4'].alignment = center_align
            ws.column_dimensions[c].width = 15
            
        ws[f'{s_col}4'].fill = header_fill_saldo
        ws[f'{e_col}4'].fill = header_fill_entrada
        col_idx += 2

    # --- Preenchimento do Corpo ---
    ws.column_dimensions['A'].width = 12
    row_idx = 5
    for mes in meses:
        ws[f'A{row_idx}'] = mes
        ws[f'A{row_idx}'].border = thin_border
        ws[f'A{row_idx}'].alignment = center_align
        
        c_idx = 1
        for emp in EMPREENDIMENTOS:
            saldo_val = calcular_saldo_acumulado_export(mes, emp, dados, meses, saldos_iniciais)
            ent_val = next((d.get('entrada', 0) for d in dados if d['mes'] == mes and d['empreendimento'] == emp), 0)
            
            s_col = get_column_letter(c_idx + 1)
            e_col = get_column_letter(c_idx + 2)
            
            # Célula de Saldo
            ws[f'{s_col}{row_idx}'] = saldo_val
            ws[f'{s_col}{row_idx}'].number_format = '#,##0.00'
            ws[f'{s_col}{row_idx}'].font = saldo_negativo_font if saldo_val < 0 else saldo_font
            ws[f'{s_col}{row_idx}'].fill = saldo_negativo_fill if saldo_val < 0 else saldo_positivo_fill
            ws[f'{s_col}{row_idx}'].border = thin_border
            ws[f'{s_col}{row_idx}'].alignment = right_align
            
            # Célula de Entrada
            ws[f'{e_col}{row_idx}'] = ent_val
            ws[f'{e_col}{row_idx}'].number_format = '#,##0.00'
            ws[f'{e_col}{row_idx}'].font = entrada_font
            ws[f'{e_col}{row_idx}'].fill = entrada_fill
            ws[f'{e_col}{row_idx}'].border = thin_border
            ws[f'{e_col}{row_idx}'].alignment = right_align
            
            c_idx += 2
        row_idx += 1

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()